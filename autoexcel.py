import openpyxl as px
import glob

# ====== (1) ファイル名から情報を抜き出す ======

# フォルダ内にあるファイル名を取得
files = glob.glob("*交通費請求明細*.xlsx")
for file in files:
    file_name = file

#import_sheet_name = '交通費_4月'
#export_file_path = 'YYYY年MM月度_交通費一覧.xlsx'
#export_sheet_name = '一覧'

# Pandasのread_excel関数でExcelファイルを読み込む
#df_order = pd.read_excel(import_file, sheet_name=import_file_ws)


# 参照先のファイル・シート指定
import_file = px.load_workbook(file_name, keep_vba=True)
import_file_ws = import_file.worksheets[2]

# コピーするファイル・シート指定
wb = px.load_workbook('YYYY年MM月度_交通費一覧.xlsx')
ws = wb.worksheets[0]


l = file_name

#前の要らない箇所以外を抜き取る
l = l[13:]
#print(l)

#後ろの.xlsxを削除
l = l[:-5]
#print(l)

#社員番号取得
employee_num = l[:3]
#print(employee_num)

#社員名取得
a = l[4:]
employee_name = a[:-7]
#print(employee_name)

#日付取得
date = l[-6:]
#print(date)

#ファイル名の情報をシートにうつす
ws.cell(row=2, column=1, value=date)
ws.cell(row=2, column=2, value=employee_num)
ws.cell(row=2, column=3, value=employee_name)


# ====== (2) ファイル内の情報を他ファイルへコピーする ======

# （19,1) 月日をコピー
cell_value = import_file_ws.cell(row=19, column=1).value
ws.cell(row=2, column=4, value=cell_value)

# （19,3）From→toをコピー
cell_value = import_file_ws.cell(row=19, column=3).value
ws.cell(row=2, column=5, value=cell_value)

# (19,4) 事由をコピー（結合セル）
cell_value = import_file_ws.cell(row=19, column=4).value
ws.cell(row=2, column=6, value=cell_value)

# （19,6）から（19,9）までコピー
for i in range(6,10):
    cell_value = import_file_ws.cell(row=19, column=i).value
    ws.cell(row=2, column=i + 1, value=cell_value)

# (19,10) 金額をコピー（結合セル）
cell_value = import_file_ws.cell(row=19, column=10).value
ws.cell(row=2, column=11, value=cell_value)


wb.save('YYYY年MM月度_交通費一覧.xlsx')
