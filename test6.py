import openpyxl as px
import re
import glob

# フォルダ内にあるファイル名を取得
files = glob.glob("*交通費請求明細*.xlsx")
for file in files:
    file_name = file

# 参照先のファイル・シート指定
exl_1 = px.load_workbook(file_name, keep_vba=True)
exl_1ws = exl_1.worksheets[2]

# コピーするファイル・シート指定
wb = px.load_workbook('YYYY年MM月度_交通費一覧.xlsx')
ws = wb.worksheets[0]


l = file_name

#前の要らない箇所以外を抜き取る
l = l[13:]
#print(l)

#後ろの.xlsxを削除
l = l[:-5]
#print(l)

#社員番号取得
employee_num = l[:3]
#print(employee_num)

#社員名取得
a = l[4:]
employee_name = a[:-7]
#print(employee_name)

#日付取得
date = l[-6:]
#print(date)

#ファイル名の情報をシートにうつす
ws.cell(row=2, column=1, value=date)
ws.cell(row=2, column=2, value=employee_num)
ws.cell(row=2, column=3, value=employee_name)

# （19,1)
cell_value = exl_1ws.cell(row=19, column=1).value
ws.cell(row=2, column=4, value=cell_value)

# （19,3）から（19,9）まで
for i in range(3, 11):
    cell_value = exl_1ws.cell(row=19, column=i).value
    ws.cell(row=2, column=i + 2, value=cell_value)

wb.save('YYYY年MM月度_交通費一覧.xlsx')
